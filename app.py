# -*- coding: utf-8 -*-
from pathlib import Path
from datetime import date, datetime
import io
import os
import re
import sqlite3
from sqlalchemy import create_engine, text as sqltext
from sqlalchemy.engine import URL
import zipfile
import html
import hashlib
import hmac
import secrets
import shutil

import altair as alt
import pandas as pd
import streamlit as st
from openpyxl import load_workbook

APP_DIR = Path(__file__).resolve().parent
DB_PATH = APP_DIR / "_local_fallback_unused.db"

st.set_page_config(
    page_title="내 자산관리 V7 Cloud Fast Mobile",
    page_icon="💰",
    layout="wide",
    initial_sidebar_state="expanded",
)

# =========================================================
# UI STYLE
# =========================================================
st.markdown("""
<style>
.stApp { background:#F6F8FB; }
header[data-testid="stHeader"] { background:rgba(246,248,251,.97); }
[data-testid="stSidebar"] { background:#FFFFFF; border-right:1px solid #E5EAF0; }
.main .block-container { max-width:1500px; padding-top:1.35rem; padding-bottom:3rem; }

.hero {
  background:linear-gradient(135deg,#173B67 0%,#2A65A0 100%);
  color:white; border-radius:19px; padding:22px 25px; margin-bottom:18px;
  box-shadow:0 8px 26px rgba(22,55,90,.13);
}
.hero h1 { color:white; margin:0; font-size:1.85rem; }
.hero p { opacity:.87; margin:7px 0 0; }

.kpi-grid {
  display:grid; grid-template-columns:repeat(4,minmax(0,1fr));
  gap:12px; margin:8px 0 19px;
}
.kpi {
  background:#fff; border:1px solid #E5EAF0; border-radius:16px;
  padding:16px; box-shadow:0 3px 12px rgba(33,54,79,.05);
}
.kpi .label { color:#687486; font-size:.85rem; font-weight:700; }
.kpi .value { color:#172333; font-size:1.42rem; font-weight:850; margin-top:7px; }
.kpi .sub { color:#8993A1; font-size:.78rem; margin-top:5px; }

.section-title { font-size:1.18rem; font-weight:850; color:#1D2A39; margin:16px 0 9px; }
.notice {
  background:#EDF6FF; color:#28597F; border:1px solid #D7E9FC;
  border-radius:13px; padding:13px 15px; margin-bottom:14px;
}
.successbox {
  background:#ECF8F0; color:#2D6A43; border:1px solid #D5EEDD;
  border-radius:13px; padding:13px 15px; margin-bottom:14px;
}
.tip {
  background:#FFF7E6; color:#80601F; border:1px solid #F2DFB5;
  border-radius:11px; padding:10px 12px; margin:8px 0 12px;
}
.account-title { font-weight:850; font-size:1.07rem; margin-bottom:6px; }
.calcbox {
  background:#F4F8FF; border:1px solid #D9E8FF; border-radius:10px;
  padding:9px 10px; margin-top:8px;
}
.flowbox {
  background:#F0FAF2; border:1px solid #DCEEDC; border-radius:10px;
  padding:8px 10px; margin:7px 0;
}
.totalbox {
  background:#FFF7EA; border:1px solid #F4E0BB; border-radius:10px;
  padding:9px 10px; margin-top:8px; font-weight:800;
}
.small { color:#7B8695; font-size:.82rem; }

div[data-testid="stDataFrame"] { border:1px solid #E4E9EF; border-radius:13px; overflow:hidden; }
.stButton>button, .stDownloadButton>button { border-radius:11px; font-weight:750; min-height:42px; }

@media(max-width:900px){
 .kpi-grid { grid-template-columns:repeat(2,minmax(0,1fr)); }
 .main .block-container { padding-left:.7rem; padding-right:.7rem; }
}
@media(max-width:600px){
 .kpi .value { font-size:1.08rem; }
 .hero h1 { font-size:1.4rem; }
}
</style>
""", unsafe_allow_html=True)

# =========================================================
# ACCOUNT STRUCTURE
# 개인1-계는 ISA + 다올 자동 합계이며 별도 원금/현재가 입력을 하지 않습니다.
# 개인1-한투는 요청에 따라 완전히 제거했습니다.
# =========================================================
EDITABLE_COMPONENTS = [
    ("보험금", "공동(보험)"),
    ("급여", "공동(급여)"),
    ("개인1-ISA", "개인"),
    ("개인1-다올", "개인"),
    ("개인2-연금", "개인(연금)"),
    ("개인2-XRP", "개인(XRP)"),
    ("오주영-증여", "증여"),
    ("오주영-용돈", "증여_용돈"),
]
COMPONENT_TO_GROUP = dict(EDITABLE_COMPONENTS)

GROUPS = [
    "개인(연금)", "개인", "개인(XRP)",
    "공동(급여)", "공동(보험)", "증여", "증여_용돈"
]

FLOW_TYPES = [
    "선택 안 함",
    "투자금(+)",
    "예수금(+)",
    "배당금(+)",
    "기타수입(+)",
    "투자금 회수(-)",
]

# =========================================================
# CLOUD DB (SUPABASE POSTGRESQL)
# =========================================================
def database_config():
    """Read Supabase connection fields separately from Streamlit Secrets."""
    try:
        cfg = st.secrets["database"]
        return {
            "host": str(cfg["host"]).strip(),
            "port": int(cfg.get("port", 6543)),
            "database": str(cfg.get("database", "postgres")).strip(),
            "user": str(cfg["user"]).strip(),
            "password": str(cfg["password"]),
        }
    except Exception as exc:
        raise RuntimeError(
            "Streamlit Secrets의 [database]에 host, port, database, user, password를 "
            "각각 설정해 주세요."
        ) from exc

@st.cache_resource
def engine():
    cfg = database_config()

    # URL.create()가 비밀번호의 !, @, #, %, :, / 등 특수문자를 안전하게 처리합니다.
    # 따라서 Secrets의 password에는 실제 비밀번호를 변환하지 않고 그대로 입력합니다.
    url = URL.create(
        drivername="postgresql+psycopg2",
        username=cfg["user"],
        password=cfg["password"],
        host=cfg["host"],
        port=cfg["port"],
        database=cfg["database"],
    )

    return create_engine(
        url,
        pool_pre_ping=True,
        pool_recycle=600,
        pool_size=3,
        max_overflow=2,
        pool_timeout=10,
        connect_args={
            "sslmode": "require",
            "connect_timeout": 8,
            "application_name": "asset_manager_v7",
        },
    )

@st.cache_resource(show_spinner=False)
def ensure_cloud_schema():
    ddl = [
        """CREATE TABLE IF NOT EXISTS settings(key TEXT PRIMARY KEY,value TEXT)""",
        """CREATE TABLE IF NOT EXISTS component_principals(component_name TEXT PRIMARY KEY,invested DOUBLE PRECISION DEFAULT 0)""",
        """CREATE TABLE IF NOT EXISTS account_meta(account_name TEXT PRIMARY KEY,start_date TEXT)""",
        """CREATE TABLE IF NOT EXISTS snapshots(id BIGSERIAL PRIMARY KEY,record_date DATE NOT NULL UNIQUE,source TEXT DEFAULT 'manual',note TEXT DEFAULT '',total_current DOUBLE PRECISION DEFAULT 0,total_invested DOUBLE PRECISION DEFAULT 0,total_profit DOUBLE PRECISION DEFAULT 0,total_return DOUBLE PRECISION DEFAULT 0,created_at TIMESTAMPTZ DEFAULT NOW())""",
        """CREATE TABLE IF NOT EXISTS component_holdings(snapshot_id BIGINT NOT NULL,component_name TEXT NOT NULL,current_value DOUBLE PRECISION DEFAULT 0,PRIMARY KEY(snapshot_id,component_name))""",
        """CREATE TABLE IF NOT EXISTS account_snapshots(snapshot_id BIGINT NOT NULL,account_name TEXT NOT NULL,invested DOUBLE PRECISION DEFAULT 0,current_value DOUBLE PRECISION DEFAULT 0,profit DOUBLE PRECISION DEFAULT 0,return_rate DOUBLE PRECISION DEFAULT 0,PRIMARY KEY(snapshot_id,account_name))""",
        """CREATE TABLE IF NOT EXISTS money_flows(id BIGSERIAL PRIMARY KEY,record_date DATE NOT NULL,component_name TEXT NOT NULL,flow_type TEXT NOT NULL,amount DOUBLE PRECISION DEFAULT 0,note TEXT DEFAULT '',principal_after DOUBLE PRECISION DEFAULT 0,created_at TIMESTAMPTZ DEFAULT NOW())""",
        """CREATE TABLE IF NOT EXISTS trend_history(record_date DATE PRIMARY KEY,return_rate DOUBLE PRECISION,invested_total DOUBLE PRECISION,source TEXT DEFAULT 'excel')""",
        """CREATE TABLE IF NOT EXISTS hlb_memo(record_date DATE PRIMARY KEY,close_price DOUBLE PRECISION DEFAULT 0,action TEXT DEFAULT '홀딩',thought TEXT DEFAULT '',created_at TIMESTAMPTZ DEFAULT NOW())""",
        """CREATE TABLE IF NOT EXISTS pension_history(record_month TEXT PRIMARY KEY,lump_sum DOUBLE PRECISION DEFAULT 0,monthly_pension DOUBLE PRECISION DEFAULT 0,retirement_allowance DOUBLE PRECISION DEFAULT 0,note TEXT DEFAULT '',created_at TIMESTAMPTZ DEFAULT NOW())""",
    ]
    with engine().begin() as conn:
        for q in ddl:
            conn.execute(sqltext(q))
    return True

@st.cache_data(ttl=60, show_spinner=False)
def sget(key, default=None):
    ensure_cloud_schema()
    with engine().connect() as conn:
        row=conn.execute(sqltext("SELECT value FROM settings WHERE key=:key"),{"key":key}).fetchone()
    return row[0] if row else default

def clear_read_cache():
    """DB 변경 후 조회 캐시를 즉시 비워 모든 기기에서 최신값을 다시 읽습니다."""
    st.cache_data.clear()

def sset(key,value):
    ensure_cloud_schema()
    with engine().begin() as conn:
        conn.execute(sqltext("INSERT INTO settings(key,value) VALUES(:key,:value) ON CONFLICT(key) DO UPDATE SET value=EXCLUDED.value"),{"key":key,"value":str(value)})
    clear_read_cache()

def initialized(): return sget("initialized","0")=="1"

@st.cache_data(ttl=30, show_spinner=False)
def principals():
    out={name:0.0 for name,_ in EDITABLE_COMPONENTS}
    with engine().connect() as conn:
        rows=conn.execute(sqltext("SELECT component_name,invested FROM component_principals")).fetchall()
    for name,value in rows:
        if name in out: out[name]=float(value or 0)
    return out

def set_principal(name,value):
    with engine().begin() as conn:
        conn.execute(sqltext("INSERT INTO component_principals(component_name,invested) VALUES(:name,:value) ON CONFLICT(component_name) DO UPDATE SET invested=EXCLUDED.invested"),{"name":name,"value":float(value or 0)})
    clear_read_cache()

def add_principal(name,delta):
    with engine().begin() as conn:
        conn.execute(sqltext("INSERT INTO component_principals(component_name,invested) VALUES(:name,:delta) ON CONFLICT(component_name) DO UPDATE SET invested=component_principals.invested+EXCLUDED.invested"),{"name":name,"delta":float(delta or 0)})
    clear_read_cache()

@st.cache_data(ttl=300, show_spinner=False)
def account_start_map():
    with engine().connect() as conn:
        rows=conn.execute(sqltext("SELECT account_name,start_date FROM account_meta")).fetchall()
    return {name:(value or '') for name,value in rows}

def save_account_starts(items):
    with engine().begin() as conn:
        for name,start_date in items.items():
            conn.execute(sqltext("INSERT INTO account_meta(account_name,start_date) VALUES(:name,:start_date) ON CONFLICT(account_name) DO UPDATE SET start_date=EXCLUDED.start_date"),{"name":name,"start_date":start_date})
    clear_read_cache()

def _read_df(query, params=None):
    with engine().connect() as conn: return pd.read_sql_query(sqltext(query),conn,params=params or {})

@st.cache_data(ttl=30, show_spinner=False)
def load_snapshots():
    d=_read_df("SELECT * FROM snapshots ORDER BY record_date,id")
    if not d.empty: d["record_date"]=pd.to_datetime(d["record_date"])
    return d

@st.cache_data(ttl=30, show_spinner=False)
def latest_snapshot_id():
    with engine().connect() as conn:
        row=conn.execute(sqltext("SELECT id FROM snapshots ORDER BY record_date DESC,id DESC LIMIT 1")).fetchone()
    return row[0] if row else None

@st.cache_data(ttl=30, show_spinner=False)
def load_component_holdings(snapshot_id):
    out={name:0.0 for name,_ in EDITABLE_COMPONENTS}
    if snapshot_id is None: return out
    with engine().connect() as conn:
        rows=conn.execute(sqltext("SELECT component_name,current_value FROM component_holdings WHERE snapshot_id=:id"),{"id":int(snapshot_id)}).fetchall()
    for name,value in rows:
        if name in out: out[name]=float(value or 0)
    return out

def latest_holdings(): return load_component_holdings(latest_snapshot_id())

@st.cache_data(ttl=30, show_spinner=False)
def load_accounts(snapshot_id=None):
    if snapshot_id is None: snapshot_id=latest_snapshot_id()
    cols=["account_name","invested","current_value","profit","return_rate"]
    if snapshot_id is None: return pd.DataFrame(columns=cols)
    order="CASE account_name WHEN '총계' THEN 1 WHEN '증여제외' THEN 2 WHEN '공동투자' THEN 3 WHEN '개인(연금)' THEN 4 WHEN '개인' THEN 5 WHEN '개인(XRP)' THEN 6 WHEN '공동(급여)' THEN 7 WHEN '공동(보험)' THEN 8 WHEN '증여' THEN 9 WHEN '증여_용돈' THEN 10 ELSE 99 END"
    return _read_df(f"SELECT account_name,invested,current_value,profit,return_rate FROM account_snapshots WHERE snapshot_id=:id ORDER BY {order}",{"id":int(snapshot_id)})

@st.cache_data(ttl=30, show_spinner=False)
def load_flows(component_name=None):
    if component_name:
        d=_read_df("SELECT * FROM money_flows WHERE component_name=:name ORDER BY record_date,id",{"name":component_name})
    else: d=_read_df("SELECT * FROM money_flows ORDER BY record_date,id")
    if not d.empty: d["record_date"]=pd.to_datetime(d["record_date"])
    return d

@st.cache_data(ttl=60, show_spinner=False)
def load_trends():
    d=_read_df("SELECT * FROM trend_history ORDER BY record_date")
    if not d.empty: d["record_date"]=pd.to_datetime(d["record_date"])
    return d

def upsert_trend(record_date,return_rate=None,invested_total=None,source="app"):
    with engine().begin() as conn:
        old=conn.execute(sqltext("SELECT return_rate,invested_total FROM trend_history WHERE record_date=:d"),{"d":record_date}).fetchone()
        rr=old[0] if old else None; invested=old[1] if old else None
        if return_rate is not None: rr=float(return_rate)
        if invested_total is not None: invested=float(invested_total)
        conn.execute(sqltext("INSERT INTO trend_history(record_date,return_rate,invested_total,source) VALUES(:d,:rr,:invested,:source) ON CONFLICT(record_date) DO UPDATE SET return_rate=EXCLUDED.return_rate,invested_total=EXCLUDED.invested_total,source=EXCLUDED.source"),{"d":record_date,"rr":rr,"invested":invested,"source":source})
    clear_read_cache()


# =========================================================
# LOGIN SECURITY
# =========================================================
def _password_hash(password, salt_hex=None):
    if salt_hex is None:
        salt = secrets.token_bytes(16)
        salt_hex = salt.hex()
    else:
        salt = bytes.fromhex(salt_hex)
    digest = hashlib.pbkdf2_hmac(
        "sha256",
        password.encode("utf-8"),
        salt,
        200_000
    )
    return salt_hex, digest.hex()

def password_is_configured():
    return bool(sget("auth_password_hash","")) and bool(sget("auth_password_salt",""))

def set_login_password(password):
    salt_hex, digest_hex = _password_hash(password)
    sset("auth_password_salt", salt_hex)
    sset("auth_password_hash", digest_hex)

def verify_login_password(password):
    salt_hex = sget("auth_password_salt","")
    saved = sget("auth_password_hash","")
    if not salt_hex or not saved:
        return False
    try:
        _, candidate = _password_hash(password, salt_hex)
        return hmac.compare_digest(saved, candidate)
    except Exception:
        return False

def auth_gate():
    if "authenticated" not in st.session_state:
        st.session_state.authenticated = False

    if st.session_state.authenticated:
        return

    st.markdown("""
    <div style="
      max-width:520px;margin:10vh auto 0 auto;background:white;
      border:1px solid #E4E9EF;border-radius:20px;padding:28px 30px;
      box-shadow:0 10px 35px rgba(28,48,72,.10);">
      <div style="font-size:1.8rem;font-weight:850;color:#172333;margin-bottom:6px;">
        🔐 내 자산관리 로그인
      </div>
      <div style="color:#6F7A89;font-size:.92rem;">
        비밀번호 인증 후 자산관리 데이터에 접근할 수 있습니다.
      </div>
    </div>
    """, unsafe_allow_html=True)

    configured = password_is_configured()

    if not configured:
        st.markdown(
            '<div style="max-width:520px;margin:18px auto 0 auto;">',
            unsafe_allow_html=True
        )
        st.info("최초 사용을 위해 4자리 이상의 숫자로 비밀번호를 설정해 주세요.")
        with st.form("initial_password_form"):
            p1 = st.text_input(
                "새 비밀번호 (4자리 이상 숫자)",
                type="password",
                max_chars=20
            )
            p2 = st.text_input(
                "새 비밀번호 확인",
                type="password",
                max_chars=20
            )
            create = st.form_submit_button(
                "비밀번호 설정",
                type="primary",
                width="stretch"
            )
        st.markdown("</div>", unsafe_allow_html=True)

        if create:
            # Streamlit 입력값을 문자열로 정규화한 뒤 PIN을 검사합니다.
            # 일부 브라우저/모바일 환경에서 보이지 않는 공백 문자가 섞이는 경우도 제거합니다.
            pin1 = re.sub(r"\s+", "", str(p1 or ""))
            pin2 = re.sub(r"\s+", "", str(p2 or ""))

            if len(pin1) < 4 or not pin1.isdecimal():
                st.error("비밀번호는 최소 4자리 숫자로 설정해 주세요.")
            elif pin1 != pin2:
                st.error("비밀번호가 서로 일치하지 않습니다.")
            else:
                set_login_password(pin1)
                st.session_state.authenticated = True
                st.rerun()
    else:
        st.markdown(
            '<div style="max-width:520px;margin:18px auto 0 auto;">',
            unsafe_allow_html=True
        )
        with st.form("login_form"):
            password = st.text_input(
                "비밀번호",
                type="password",
                placeholder="4자리 이상 숫자"
            )
            login = st.form_submit_button(
                "로그인",
                type="primary",
                width="stretch"
            )
        st.markdown("</div>", unsafe_allow_html=True)

        if login:
            if verify_login_password(password):
                st.session_state.authenticated = True
                st.rerun()
            else:
                st.error("비밀번호가 올바르지 않습니다.")

    st.stop()

def logout_button():
    if st.button("🔒 로그아웃", width="stretch"):
        st.session_state.authenticated = False
        st.rerun()

def change_password_ui():
    st.markdown("### 로그인 비밀번호 변경")
    with st.form("change_password_form"):
        current = st.text_input("현재 비밀번호", type="password")
        new1 = st.text_input("새 비밀번호 (4자리 이상 숫자)", type="password")
        new2 = st.text_input("새 비밀번호 확인", type="password")
        change = st.form_submit_button("비밀번호 변경", width="stretch")
    if change:
        if not verify_login_password(current):
            st.error("현재 비밀번호가 올바르지 않습니다.")
        else:
            new_pin1 = re.sub(r"\s+", "", str(new1 or ""))
            new_pin2 = re.sub(r"\s+", "", str(new2 or ""))
            if len(new_pin1) < 4 or not new_pin1.isdecimal():
                st.error("새 비밀번호는 최소 4자리 숫자로 설정해 주세요.")
            elif new_pin1 != new_pin2:
                st.error("새 비밀번호가 서로 일치하지 않습니다.")
            else:
                set_login_password(new_pin1)
            st.success("로그인 비밀번호를 변경했습니다.")

# =========================================================
# FORMAT / INPUT
# =========================================================
def num(value):
    try:
        return 0.0 if value is None else float(value)
    except Exception:
        return 0.0

def kr(value):
    value = num(value)
    sign = "-" if value < 0 else ""
    value = abs(value)
    eok = int(value // 100_000_000)
    man = int((value % 100_000_000) // 10_000)
    if eok and man:
        return f"{sign}{eok}억 {man:,}만원"
    if eok:
        return f"{sign}{eok}억원"
    if man:
        return f"{sign}{man:,}만원"
    return f"{sign}{int(value):,}원"

def pc(value):
    return f"{num(value)*100:+.1f}%"

def card(label, value, sub=""):
    return (
        '<div class="kpi">'
        f'<div class="label">{html.escape(label)}</div>'
        f'<div class="value">{html.escape(value)}</div>'
        f'<div class="sub">{html.escape(sub)}</div>'
        '</div>'
    )

def parse_money_text(value):
    s = str(value).replace(",","").replace("원","").strip()
    if not s:
        return 0.0
    try:
        return float(s)
    except Exception:
        return 0.0

def _format_money_state(key):
    value = parse_money_text(st.session_state.get(key, "0"))
    st.session_state[key] = f"{int(value):,}"

def money_input(label, default, key, help_text=None):
    if key not in st.session_state:
        st.session_state[key] = f"{int(num(default)):,}"
    value = st.text_input(
        label,
        key=key,
        on_change=_format_money_state,
        args=(key,),
        help=help_text,
        placeholder="0"
    )
    return parse_money_text(value)

def parse_date_text(value):
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    if value is None:
        return None
    s = str(value).strip().replace(" ","")
    m = re.match(r"^(\d{2,4})\.(\d{1,2})\.(\d{0,2})\.?$", s)
    if not m:
        return None
    year = int(m.group(1))
    month = int(m.group(2))
    day = int(m.group(3) or 1)
    if year < 100:
        year += 2000
    try:
        return date(year, month, day)
    except Exception:
        return None

def date_from_filename(name):
    m = re.search(r"(\d{2})\.(\d{1,2})\.(\d{1,2})", name)
    if not m:
        return None
    try:
        return date(
            2000 + int(m.group(1)),
            int(m.group(2)),
            int(m.group(3))
        )
    except Exception:
        return None


def cloud_config_status():
    """Check whether cloud DB secrets have been configured."""
    try:
        db = st.secrets.get("database", {})
        return bool(db.get("url"))
    except Exception:
        return False

def format_start_date(value):
    if not value:
        return "-"
    try:
        return pd.to_datetime(value).strftime("%Y-%m-%d")
    except Exception:
        return str(value)

# =========================================================
# PORTFOLIO CALCULATION
# =========================================================
def build_accounts(holdings, principal_override=None):
    p = principal_override if principal_override is not None else principals()

    group_current = {g: 0.0 for g in GROUPS}
    group_principal = {g: 0.0 for g in GROUPS}

    for component, group in EDITABLE_COMPONENTS:
        group_current[group] += num(holdings.get(component, 0))
        group_principal[group] += num(p.get(component, 0))

    details = []
    for group in GROUPS:
        invested = group_principal[group]
        current = group_current[group]
        profit = current - invested
        details.append({
            "account_name": group,
            "invested": invested,
            "current_value": current,
            "profit": profit,
            "return_rate": profit / invested if invested else 0,
        })

    def combine(name, names):
        invested = sum(
            x["invested"] for x in details if x["account_name"] in names
        )
        current = sum(
            x["current_value"] for x in details if x["account_name"] in names
        )
        profit = current - invested
        return {
            "account_name": name,
            "invested": invested,
            "current_value": current,
            "profit": profit,
            "return_rate": profit / invested if invested else 0,
        }

    total = combine("총계", GROUPS)
    gift_excluded = combine(
        "증여제외",
        ["개인(연금)","개인","개인(XRP)","공동(급여)","공동(보험)"]
    )
    joint = combine("공동투자", ["공동(급여)","공동(보험)"])

    return [total, gift_excluded, joint] + details

def save_summary(record_date, holdings, source, note):
    accounts=build_accounts(holdings); total=accounts[0]
    with engine().begin() as conn:
        row=conn.execute(sqltext("SELECT id FROM snapshots WHERE record_date=:d"),{"d":record_date}).fetchone()
        if row:
            sid=row[0]
            conn.execute(sqltext("UPDATE snapshots SET source=:source,note=:note,total_current=:current,total_invested=:invested,total_profit=:profit,total_return=:rr WHERE id=:id"),{"source":source,"note":note,"current":total["current_value"],"invested":total["invested"],"profit":total["profit"],"rr":total["return_rate"],"id":sid})
            conn.execute(sqltext("DELETE FROM component_holdings WHERE snapshot_id=:id"),{"id":sid})
            conn.execute(sqltext("DELETE FROM account_snapshots WHERE snapshot_id=:id"),{"id":sid})
        else:
            sid=conn.execute(sqltext("INSERT INTO snapshots(record_date,source,note,total_current,total_invested,total_profit,total_return) VALUES(:d,:source,:note,:current,:invested,:profit,:rr) RETURNING id"),{"d":record_date,"source":source,"note":note,"current":total["current_value"],"invested":total["invested"],"profit":total["profit"],"rr":total["return_rate"]}).scalar_one()
        for name,_ in EDITABLE_COMPONENTS:
            conn.execute(sqltext("INSERT INTO component_holdings(snapshot_id,component_name,current_value) VALUES(:sid,:name,:value)"),{"sid":sid,"name":name,"value":num(holdings.get(name,0))})
        for a in accounts:
            conn.execute(sqltext("INSERT INTO account_snapshots(snapshot_id,account_name,invested,current_value,profit,return_rate) VALUES(:sid,:name,:invested,:current,:profit,:rr)"),{"sid":sid,"name":a["account_name"],"invested":a["invested"],"current":a["current_value"],"profit":a["profit"],"rr":a["return_rate"]})
    upsert_trend(record_date,total["return_rate"],total["invested"],source="app")

# =========================================================
# INITIAL EXCEL IMPORT
# =========================================================
def parse_initial_excel(uploaded):
    raw = uploaded.getvalue()
    if len(raw) < 4 or raw[:2] != b"PK":
        raise ValueError("정상적인 Excel 통합문서(.xlsx) 파일이 아닙니다.")

    try:
        wb = load_workbook(
            io.BytesIO(raw),
            data_only=True,
            read_only=False
        )
    except zipfile.BadZipFile as exc:
        raise ValueError("파일이 손상되었거나 .xlsx 형식이 아닙니다.") from exc

    if "주식(총괄)" not in wb.sheetnames:
        raise ValueError("'주식(총괄)' 시트를 찾지 못했습니다.")

    ws = wb["주식(총괄)"]

    holdings = {
        "보험금": num(ws["C4"].value),
        "급여": num(ws["D4"].value),
        "개인1-ISA": num(ws["G4"].value),
        "개인1-다올": num(ws["H4"].value),
        "개인2-연금": num(ws["I4"].value),
        "개인2-XRP": num(ws["J4"].value),
        "오주영-증여": num(ws["K4"].value),
        "오주영-용돈": num(ws["L4"].value),
    }

    group_rows = {}
    start_dates = {}
    for row in range(9, 19):
        name = ws.cell(row, 1).value
        if not name:
            continue
        group_rows[str(name)] = {
            "invested": num(ws.cell(row,2).value),
            "current": num(ws.cell(row,3).value),
        }
        start = ws.cell(row,6).value
        if isinstance(start, datetime):
            start_dates[str(name)] = start.date().isoformat()
        elif isinstance(start, date):
            start_dates[str(name)] = start.isoformat()
        elif start:
            start_dates[str(name)] = str(start)
        else:
            start_dates[str(name)] = ""

    # V7 Cloud: 세부 계좌의 투자원금은 각 원본 시트의 명시된 기준값을 직접 사용합니다.
    # - 개인1-ISA  : 주식(개인)_ISA!E209  = 순투자금(입금-출금)
    # - 개인1-다올 : 주식(개인)_다올!E170 = 순투자금(입금-출금)
    # - 개인2-XRP  : 가상화폐(XRP)!E76    = 순투자금(입금-출금)
    isa_principal = 0.0
    daol_principal = 0.0
    xrp_principal = 0.0

    try:
        isa_principal = num(wb["주식(개인)_ISA"]["E209"].value)
    except Exception:
        isa_principal = 0.0

    try:
        daol_principal = num(wb["주식(개인)_다올"]["E170"].value)
    except Exception:
        daol_principal = 0.0

    try:
        xrp_principal = num(wb["가상화폐(XRP)"]["E76"].value)
    except Exception:
        xrp_principal = 0.0

    principal = {
        "보험금": group_rows.get("공동(보험)",{}).get("invested",0),
        "급여": group_rows.get("공동(급여)",{}).get("invested",0),
        "개인1-ISA": isa_principal,
        "개인1-다올": daol_principal,
        "개인2-연금": group_rows.get("개인(연금)",{}).get("invested",0),
        "개인2-XRP": xrp_principal,
        "오주영-증여": group_rows.get("증여",{}).get("invested",0),
        "오주영-용돈": group_rows.get("증여_용돈",{}).get("invested",0),
    }

    returns = []
    invested_history = []
    for col in range(15, ws.max_column + 1):
        d1 = parse_date_text(ws.cell(2,col).value)
        rr = ws.cell(3,col).value
        if d1 and rr is not None:
            returns.append((d1.isoformat(), num(rr)))

        d2 = parse_date_text(ws.cell(7,col).value)
        invested = ws.cell(8,col).value
        if d2 and invested is not None:
            invested_history.append((d2.isoformat(), num(invested)))

    historical_dates = [
        date.fromisoformat(item[0])
        for item in returns + invested_history
    ]
    reference_date = (
        date_from_filename(uploaded.name)
        or (max(historical_dates) if historical_dates else date.today())
    )

    accounts = build_accounts(holdings, principal)

    return {
        "holdings": holdings,
        "principal": principal,
        "accounts": accounts,
        "start_dates": start_dates,
        "returns": returns,
        "invested_history": invested_history,
        "reference_date": reference_date,
        "sheet_count": len(wb.sheetnames),
        "file_hash": hashlib.sha256(raw).hexdigest(),
    }

def initialize_excel(parsed, filename):
    if initialized():
        return False

    for name, value in parsed["principal"].items():
        set_principal(name, value)

    save_account_starts(parsed["start_dates"])

    for d, rr in parsed["returns"]:
        upsert_trend(d, return_rate=rr, source="excel")
    for d, invested in parsed["invested_history"]:
        upsert_trend(d, invested_total=invested, source="excel")

    save_summary(
        parsed["reference_date"].isoformat(),
        parsed["holdings"],
        "excel_initial",
        f"{filename} 최초 데이터 이관"
    )

    sset("initialized","1")
    sset("initial_file", filename)
    sset("initial_date", parsed["reference_date"].isoformat())
    sset("initial_hash", parsed["file_hash"])
    return True

# =========================================================
# TABLE / CHART
# =========================================================
def colored_table(df):
    def style(row):
        colors = {
            "총계":"background-color:#E6F1FF;font-weight:700;",
            "증여제외":"background-color:#FFF1D6;font-weight:650;",
            "공동투자":"background-color:#E4F5E7;font-weight:650;",
            "공동(급여)":"background-color:#F0E3FF;font-weight:650;",
        }
        value = colors.get(row["계좌명"], "")
        return [value] * len(row)
    return df.style.apply(style, axis=1)

def show_investment_table(accounts):
    if accounts.empty:
        st.info("투자계좌 데이터가 없습니다.")
        return

    starts = account_start_map()
    t = accounts.copy()
    t.columns = ["계좌명","순투자금","현재자산","수익금","수익률"]
    t["투자시작"] = t["계좌명"].map(
        lambda x: format_start_date(starts.get(x,""))
    )

    for col in ["순투자금","현재자산","수익금"]:
        t[col] = t[col].map(lambda x: f"{num(x):,.0f}")
    t["수익률"] = t["수익률"].map(
        lambda x: f"{num(x)*100:.1f}%"
    )

    st.dataframe(
        colored_table(t),
        width="stretch",
        hide_index=True,
        height=440
    )
    st.caption(
        "색상 구분: 총계 · 증여제외 · 공동투자 · 공동(급여)"
    )

def trend_chart(df, y, title, money=False):
    if df.empty:
        st.info(f"{title} 이력이 없습니다.")
        return

    chart_df = df[["record_date",y]].dropna().copy()
    if chart_df.empty:
        st.info(f"{title} 이력이 없습니다.")
        return

    chart_df = chart_df.drop_duplicates(
        "record_date",
        keep="last"
    )

    if y == "return_rate":
        chart_df["value"] = chart_df[y] * 100
    else:
        chart_df["value"] = chart_df[y]

    chart = (
        alt.Chart(chart_df)
        .mark_line(point=False, strokeWidth=2)
        .encode(
            x=alt.X(
                "record_date:T",
                title=None,
                axis=alt.Axis(
                    format="%Y.%m",
                    labelAngle=0,
                    tickCount=8
                )
            ),
            y=alt.Y(
                "value:Q",
                title=None,
                axis=alt.Axis(format="~s") if money else alt.Axis()
            ),
            tooltip=[
                alt.Tooltip(
                    "record_date:T",
                    title="날짜",
                    format="%Y.%m.%d"
                ),
                alt.Tooltip(
                    "value:Q",
                    title=title,
                    format=",.0f" if money else ",.1f"
                ),
            ]
        )
        .properties(height=215)
    )
    st.altair_chart(chart, width="stretch")

@st.cache_data(ttl=30, show_spinner=False)
def current_accounts_live():
    accounts = build_accounts(latest_holdings())
    return pd.DataFrame(accounts)[
        ["account_name","invested","current_value","profit","return_rate"]
    ]

def render_investment_dashboard():
    left, right = st.columns([1.12,.88], gap="medium")

    with left:
        st.markdown("#### 현재 투자 현황")
        show_investment_table(current_accounts_live())

    trends = load_trends()
    with right:
        st.markdown("#### 전체 투자수익률 추세")
        trend_chart(trends, "return_rate", "수익률")
        rr = trends.dropna(subset=["return_rate"]) if not trends.empty else pd.DataFrame()
        if not rr.empty:
            st.caption(
                f"최근 수익률: {num(rr.iloc[-1]['return_rate'])*100:+.1f}%"
            )

        st.markdown("#### 총 투자금액 추이")
        trend_chart(trends, "invested_total", "총 투자금액", money=True)
        ii = trends.dropna(subset=["invested_total"]) if not trends.empty else pd.DataFrame()
        if not ii.empty:
            st.caption(
                f"최근 총 투자금액: {kr(ii.iloc[-1]['invested_total'])}"
            )

# =========================================================
# INVESTMENT PRINCIPAL TRANSACTIONS
# =========================================================
def principal_after_for_component(component):
    return num(principals().get(component,0))

def save_flow(component, record_date, flow_type, amount, note):
    amount=num(amount)
    if flow_type=="선택 안 함" or amount<=0: return False
    delta=amount if flow_type in ("투자금(+)","예수금(+)","배당금(+)","기타수입(+)") else (-amount if flow_type=="투자금 회수(-)" else 0)
    if delta: add_principal(component,delta)
    principal_after=principal_after_for_component(component)
    with engine().begin() as conn:
        conn.execute(sqltext("INSERT INTO money_flows(record_date,component_name,flow_type,amount,note,principal_after) VALUES(:d,:component,:flow_type,:amount,:note,:principal_after)"),{"d":record_date,"component":component,"flow_type":flow_type,"amount":amount,"note":note,"principal_after":principal_after})
    if delta: upsert_trend(record_date,invested_total=sum(principals().values()),source="app")
    return True

def flow_history_table(component):
    history = load_flows(component)
    if history.empty:
        st.caption("저장된 순투자이력이 없습니다.")
        return

    show = history.copy()
    show["일자"] = show["record_date"].dt.strftime("%Y-%m-%d")
    show["금액"] = show["amount"].map(
        lambda x: f"{num(x):,.0f}원"
    )
    show["반영 후 순투자금"] = show["principal_after"].map(
        lambda x: f"{num(x):,.0f}원"
    )
    show = show[
        ["일자","flow_type","금액","note","반영 후 순투자금"]
    ].rename(columns={
        "flow_type":"내역",
        "note":"메모"
    })

    st.dataframe(
        show,
        width="stretch",
        hide_index=True,
        height=min(320, 38 * (len(show)+1))
    )

# =========================================================
# HLB
# =========================================================
@st.cache_data(ttl=30, show_spinner=False)
def load_hlb():
    d=_read_df("SELECT * FROM hlb_memo ORDER BY record_date")
    if not d.empty: d["record_date"]=pd.to_datetime(d["record_date"])
    return d

def save_hlb(record_date,close_price,action,thought):
    with engine().begin() as conn:
        conn.execute(sqltext("INSERT INTO hlb_memo(record_date,close_price,action,thought) VALUES(:d,:price,:action,:thought) ON CONFLICT(record_date) DO UPDATE SET close_price=EXCLUDED.close_price,action=EXCLUDED.action,thought=EXCLUDED.thought"),{"d":record_date,"price":float(close_price),"action":action,"thought":thought})

# =========================================================
# PENSION
# =========================================================
@st.cache_data(ttl=30, show_spinner=False)
def load_pension():
    return _read_df("SELECT * FROM pension_history ORDER BY record_month")

def validate_year_month(value):
    s=str(value).strip(); m=re.match(r"^(\d{4})[.\-/](\d{1,2})$",s)
    if not m: return None
    year=int(m.group(1)); month=int(m.group(2))
    if not 1<=month<=12: return None
    return f"{year:04d}.{month:02d}"

def save_pension(record_month,lump_sum,monthly_pension,retirement_allowance,note):
    with engine().begin() as conn:
        conn.execute(sqltext("INSERT INTO pension_history(record_month,lump_sum,monthly_pension,retirement_allowance,note) VALUES(:ym,:lump,:monthly,:retire,:note) ON CONFLICT(record_month) DO UPDATE SET lump_sum=EXCLUDED.lump_sum,monthly_pension=EXCLUDED.monthly_pension,retirement_allowance=EXCLUDED.retirement_allowance,note=EXCLUDED.note"),{"ym":record_month,"lump":float(lump_sum),"monthly":float(monthly_pension),"retire":float(retirement_allowance),"note":note})

def pension_history_view():
    d = load_pension()
    if d.empty:
        st.info("아직 저장된 연금내역이 없습니다.")
        return

    d = d.sort_values("record_month").reset_index(drop=True)

    for source, prefix in [
        ("lump_sum", "일시금"),
        ("monthly_pension", "연금(월)"),
        ("retirement_allowance", "퇴직수당"),
    ]:
        change = d[source].diff()
        rate = d[source].pct_change()
        d[f"{prefix}_변동액"] = change
        d[f"{prefix}_변동률"] = rate

    rows = []
    for idx, row in d.iterrows():
        first = idx == 0
        rows.append({
            "년월": row["record_month"],
            "일시금": f"{num(row['lump_sum']):,.0f}",
            "일시금 변동액": "-" if first else f"{num(row['일시금_변동액']):+,.0f}",
            "일시금 변동률": "-" if first else f"{num(row['일시금_변동률'])*100:+.2f}%",
            "연금(월)": f"{num(row['monthly_pension']):,.0f}",
            "연금 변동액": "-" if first else f"{num(row['연금(월)_변동액']):+,.0f}",
            "연금 변동률": "-" if first else f"{num(row['연금(월)_변동률'])*100:+.2f}%",
            "퇴직수당": f"{num(row['retirement_allowance']):,.0f}",
            "퇴직수당 변동액": "-" if first else f"{num(row['퇴직수당_변동액']):+,.0f}",
            "퇴직수당 변동률": "-" if first else f"{num(row['퇴직수당_변동률'])*100:+.2f}%",
            "메모": row["note"] or "",
        })

    st.dataframe(
        pd.DataFrame(rows),
        width="stretch",
        hide_index=True,
        height=430
    )

# =========================================================
# BACKUP / RESTORE
# =========================================================
def validate_sqlite_bytes(raw):
    tmp = APP_DIR / "_restore_check.db"
    tmp.write_bytes(raw)
    try:
        c = sqlite3.connect(tmp)
        c.execute("SELECT name FROM sqlite_master LIMIT 1").fetchall()
        required = c.execute(
            "SELECT name FROM sqlite_master "
            "WHERE type='table' AND name='settings'"
        ).fetchone()
        c.close()
        return required is not None
    except Exception:
        return False
    finally:
        if tmp.exists():
            tmp.unlink()

def restore_db(raw):
    stamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    safety = APP_DIR / f"auto_backup_before_restore_{stamp}.db"
    if DB_PATH.exists():
        shutil.copy2(DB_PATH, safety)
    tmp = APP_DIR / "_restore_new.db"
    tmp.write_bytes(raw)
    os.replace(tmp, DB_PATH)
    return safety

# =========================================================
# CLOUD CONNECTION CHECK
# =========================================================
try:
    ensure_cloud_schema()
except Exception as exc:
    st.error("☁️ Supabase 데이터베이스 연결에 실패했습니다.")
    st.code(str(exc))
    st.info("Streamlit Cloud → Manage app → Settings → Secrets의 [database] host / port / database / user / password를 확인해 주세요.")
    st.stop()

# =========================================================
# AUTH
# =========================================================
auth_gate()

st.markdown("""
<style>
@media (max-width: 768px) {
    .main .block-container {
        padding-top: 0.7rem;
        padding-left: 0.8rem;
        padding-right: 0.8rem;
    }
    div[data-testid="stSelectbox"] {
        margin-bottom: 0.4rem;
    }
    div[data-testid="stSelectbox"] label p {
        font-weight: 700;
        font-size: 0.95rem;
    }
}
</style>
""", unsafe_allow_html=True)

# =========================================================
# NAVIGATION / SIDEBAR
# =========================================================
init = initialized()

NAV_OPTIONS = [
    "주식(총괄)",
    "주식(기록_현재가격)",
    "주식(기록_투자원금)",
    "HLB 투자 메모",
    "연금내역",
    "기록 관리",
    "설정/백업",
]

if "nav_page" not in st.session_state:
    st.session_state["nav_page"] = "주식(총괄)"

def _sync_from_sidebar():
    selected = st.session_state.get("sidebar_nav", "주식(총괄)")
    st.session_state["nav_page"] = selected
    st.session_state["mobile_nav"] = selected

def _sync_from_mobile():
    selected = st.session_state.get("mobile_nav", "주식(총괄)")
    st.session_state["nav_page"] = selected
    st.session_state["sidebar_nav"] = selected

with st.sidebar:
    st.markdown("## 💰 내 자산관리")
    st.caption("내 자산관리 V7 Cloud Fast Mobile")

    if not init:
        page = "초기 설정"
        st.info("최초 1회 기존 Excel을 가져옵니다.")
        st.divider()
        logout_button()
    else:
        current = st.session_state.get("nav_page", "주식(총괄)")
        if st.session_state.get("sidebar_nav") not in NAV_OPTIONS:
            st.session_state["sidebar_nav"] = current

        st.radio(
            "메뉴",
            NAV_OPTIONS,
            index=NAV_OPTIONS.index(current),
            key="sidebar_nav",
            on_change=_sync_from_sidebar,
            label_visibility="collapsed",
        )

        st.divider()
        st.success("⚡ Supabase 클라우드 고속 모드")
        st.caption(f"초기 파일: {sget('initial_file','-')}")
        st.caption(f"앱 관리 시작일: {sget('initial_date','-')}")
        st.caption("이후에는 Excel 없이 계속 관리합니다.")
        st.divider()
        logout_button()

if init:
    current = st.session_state.get("nav_page", "주식(총괄)")
    if st.session_state.get("mobile_nav") not in NAV_OPTIONS:
        st.session_state["mobile_nav"] = current

    st.selectbox(
        "📱 메뉴 이동",
        NAV_OPTIONS,
        index=NAV_OPTIONS.index(current),
        key="mobile_nav",
        on_change=_sync_from_mobile,
        help="휴대전화에서는 이 메뉴를 이용하면 왼쪽 사이드바를 열지 않아도 됩니다.",
    )
    page = st.session_state.get("nav_page", "주식(총괄)")

# =========================================================
# INITIAL SETUP
# =========================================================
if not init:
    st.markdown(
        '<div class="hero"><h1>📥 최초 데이터 가져오기</h1>'
        '<p>기존 엑셀은 처음 한 번만 사용합니다. '
        '이후에는 V7 Cloud 데이터베이스만으로 관리합니다.</p></div>',
        unsafe_allow_html=True
    )

    uploaded = st.file_uploader(
        "기존 자산관리 Excel(.xlsx)",
        type=["xlsx"]
    )

    if uploaded is not None:
        try:
            parsed = parse_initial_excel(uploaded)
            total = parsed["accounts"][0]

            st.success(
                f"시트 {parsed['sheet_count']}개를 정상적으로 읽었습니다."
            )

            st.markdown(
                '<div class="kpi-grid">' +
                card("총 투자금액", kr(total["invested"]), "초기 이관") +
                card("현재 총 자산", kr(total["current_value"]), "초기 이관") +
                card("수익금", kr(total["profit"]), "초기 이관") +
                card("수익률", pc(total["return_rate"]), "초기 이관") +
                '</div>',
                unsafe_allow_html=True
            )

            st.write(
                f"**앱 관리 시작 기준일:** "
                f"{parsed['reference_date'].isoformat()}"
            )

            preview = pd.DataFrame(parsed["accounts"])[
                ["account_name","invested","current_value","profit","return_rate"]
            ]
            # 투자시작 미리보기도 DB 저장 전 임시 표시
            starts_backup = account_start_map()
            save_account_starts(parsed["start_dates"])
            show_investment_table(preview)
            # 초기 설정 상태라 DB에 start_dates가 잠깐 저장되어도 이후 초기 등록에서 동일값으로 갱신됨.
            # 사용자가 취소해도 자산값은 저장되지 않음.

            ok = st.checkbox(
                "이 파일을 V7 Cloud 최초 데이터로 등록합니다.",
                key="confirm_initial_cloud_import",
            )

            if st.button(
                "최초 데이터 등록",
                type="primary",
                width="stretch",
                disabled=not ok,
                key="initial_cloud_import_button",
            ):
                with st.status(
                    "☁️ Supabase에 최초 데이터를 등록하고 있습니다...",
                    expanded=True,
                ) as status:
                    try:
                        status.write("1/4 계좌별 투자원금을 저장합니다.")
                        initialize_excel(parsed, uploaded.name)

                        status.write("2/4 현재 자산 및 투자시작일을 저장합니다.")
                        status.write("3/4 과거 수익률·총 투자금액 추이를 저장합니다.")
                        status.write("4/4 초기화 상태를 확인합니다.")

                        if initialized():
                            status.update(
                                label="✅ 최초 데이터 등록이 완료되었습니다.",
                                state="complete",
                                expanded=True,
                            )
                            st.success(
                                "Supabase 최초 데이터 등록이 완료되었습니다. "
                                "이제 Excel 없이 클라우드에서 계속 관리할 수 있습니다."
                            )
                            st.session_state["initial_import_done"] = True
                        else:
                            status.update(
                                label="⚠️ 등록 결과를 확인하지 못했습니다.",
                                state="error",
                                expanded=True,
                            )
                            st.error(
                                "저장 함수는 실행됐지만 초기화 완료 상태가 확인되지 않았습니다."
                            )

                    except Exception as exc:
                        status.update(
                            label="❌ 최초 데이터 등록 중 오류가 발생했습니다.",
                            state="error",
                            expanded=True,
                        )
                        st.exception(exc)

        except Exception as exc:
            st.error(f"가져오기 실패: {exc}")

    if st.session_state.get("initial_import_done"):
        if st.button(
            "주식(총괄)로 이동",
            type="primary",
            width="stretch",
            key="go_dashboard_after_initial_import",
        ):
            st.session_state.pop("initial_import_done", None)
            st.rerun()

    st.stop()

# =========================================================
# NORMAL PAGES
# =========================================================
snaps = load_snapshots()

if page == "주식(총괄)":
    st.markdown(
        '<div class="hero"><h1>내 주식자산 한눈에 보기</h1>'
        '<p>현재 주식자산과 장기 수익률·투자금 추이를 '
        '앱 데이터로 관리합니다.</p></div>',
        unsafe_allow_html=True
    )

    live = build_accounts(latest_holdings())[0]

    st.markdown(
        '<div class="kpi-grid">' +
        card("총 투자금액", kr(live["invested"]), "투자원금 기준") +
        card("현재 총 자산", kr(live["current_value"]), "최근 현재가격 기준") +
        card("수익금", kr(live["profit"]), "현재 총 자산 - 총 투자금액") +
        card("수익률", pc(live["return_rate"]), "총계 기준") +
        '</div>',
        unsafe_allow_html=True
    )

    st.markdown(
        '<div class="section-title">투자계좌 요약</div>',
        unsafe_allow_html=True
    )
    render_investment_dashboard()

elif page == "투자계좌":
    st.markdown("## 📊 투자계좌")
    render_investment_dashboard()

elif page == "주식(기록_현재가격)":
    st.markdown("## ✍️ 주식(기록_현재가격)")
    st.caption(
        "직전 저장값이 자동으로 채워집니다. "
        "변경된 현재가격만 수정한 뒤 저장하세요."
    )

    defaults = latest_holdings()

    d = st.date_input("기준일", date.today())

    cols = st.columns(2)
    values = {}

    # 개인1-계는 자동합계이므로 먼저 ISA/다올을 입력
    ordered = [
        "보험금","급여",
        "개인1-ISA","개인1-다올",
        "개인2-연금","개인2-XRP",
        "오주영-증여","오주영-용돈"
    ]

    for idx, name in enumerate(ordered):
        with cols[idx % 2]:
            values[name] = money_input(
                name,
                defaults.get(name,0),
                f"current_{name}"
            )

    personal_total = (
        values["개인1-ISA"] +
        values["개인1-다올"]
    )
    st.info(
        f"개인1-계(자동합계): **{personal_total:,.0f}원** "
        f"= 개인1-ISA + 개인1-다올"
    )

    preview = build_accounts(values)
    total = preview[0]

    st.markdown(
        '<div class="kpi-grid">' +
        card("총 투자금액", kr(total["invested"]), "저장 전 미리보기") +
        card("현재 총 자산", kr(total["current_value"]), "저장 전 미리보기") +
        card("수익금", kr(total["profit"]), "저장 전 미리보기") +
        card("수익률", pc(total["return_rate"]), "저장 전 미리보기") +
        '</div>',
        unsafe_allow_html=True
    )

    st.markdown(
        '<div class="tip">💡 금액 입력 후 Enter를 누르거나 다른 입력칸으로 '
        '이동하면 입력창에 천 단위 쉼표가 적용됩니다. '
        '저장 전 값은 대시보드와 그래프에 반영되지 않습니다.</div>',
        unsafe_allow_html=True
    )

    note = st.text_input(
        "메모",
        placeholder="예: 월말 평가금액"
    )

    if st.button(
        "현재가격 기록 저장",
        type="primary",
        width="stretch"
    ):
        save_summary(
            d.isoformat(),
            values,
            "manual_current_price",
            note
        )
        st.success("현재가격 기록을 저장했습니다.")
        st.rerun()

elif page == "주식(기록_투자원금)":
    st.markdown("## 💵 주식(기록_투자원금)")
    st.caption(
        "모든 계좌의 기존 투자금과 거래이력을 한 화면에서 확인합니다. "
        "변동이 있는 계좌만 내역·일자·금액·메모를 입력하세요."
    )

    p = principals()

    st.info(
        f"개인1-계(자동합계) 기존 투자금: "
        f"**{p.get('개인1-ISA',0)+p.get('개인1-다올',0):,.0f}원** "
        f"= 개인1-ISA + 개인1-다올"
    )

    pending = {}

    for start in range(0, len(EDITABLE_COMPONENTS), 3):
        columns = st.columns(3, gap="small")

        for offset, (name, _) in enumerate(
            EDITABLE_COMPONENTS[start:start+3]
        ):
            with columns[offset]:
                with st.container(border=True):
                    existing = num(p.get(name,0))

                    st.markdown(
                        f'<div class="account-title">'
                        f'{html.escape(name)}</div>',
                        unsafe_allow_html=True
                    )

                    st.markdown(
                        '<div class="calcbox">'
                        '<b>1. 투자금 내역(기존)</b><br>'
                        f'기존 투자금&nbsp;&nbsp;<b>{existing:,.0f}원</b>'
                        '</div>',
                        unsafe_allow_html=True
                    )

                    st.markdown(
                        '<div class="flowbox"><b>2. 추가투입 내역</b></div>',
                        unsafe_allow_html=True
                    )

                    flow_date = st.date_input(
                        "거래일자",
                        date.today(),
                        key=f"flow_date_{name}"
                    )

                    flow_type = st.selectbox(
                        "내역 선택",
                        FLOW_TYPES,
                        key=f"flow_type_{name}"
                    )

                    amount = money_input(
                        "금액",
                        0,
                        f"flow_amount_{name}"
                    )

                    memo = st.text_input(
                        "메모",
                        key=f"flow_memo_{name}",
                        placeholder="특이사항"
                    )

                    projected = existing
                    if flow_type in ("투자금(+)", "예수금(+)", "배당금(+)", "기타수입(+)"):
                        projected += amount
                    elif flow_type == "투자금 회수(-)":
                        projected = max(0, projected - amount)

                    st.markdown(
                        '<div class="totalbox">'
                        '<b>3. 총 투자금계</b><br>'
                        f'총 투자금액&nbsp;&nbsp;'
                        f'<span style="color:#1459C7">'
                        f'{projected:,.0f}원</span>'
                        '</div>',
                        unsafe_allow_html=True
                    )

                    with st.expander("순투자이력"):
                        flow_history_table(name)

                    pending[name] = {
                        "date": flow_date,
                        "type": flow_type,
                        "amount": amount,
                        "memo": memo,
                    }

    existing_total = sum(p.values())
    delta_total = 0.0
    for item in pending.values():
        if item["type"] in ("투자금(+)", "예수금(+)", "배당금(+)", "기타수입(+)"):
            delta_total += item["amount"]
        elif item["type"] == "투자금 회수(-)":
            delta_total -= item["amount"]

    st.markdown(
        '<div class="kpi-grid">' +
        card("기존 총 투자금액", kr(existing_total), "저장 전") +
        card(
            "투자금 변동",
            kr(delta_total),
            "투자금·예수금·배당금·기타수입(+) - 투자금 회수(-)"
        ) +
        card(
            "저장 후 총 투자금액",
            kr(max(0,existing_total+delta_total)),
            "예상값"
        ) +
        card(
            "입력 계좌 수",
            f"{sum(1 for x in pending.values() if x['type']!='선택 안 함' and x['amount']>0)}개",
            "변동 있는 계좌"
        ) +
        '</div>',
        unsafe_allow_html=True
    )

    st.markdown(
        '<div class="tip">💡 예수금(+), 배당금(+), 기타수입(+)은 '
        '거래이력에 저장되지만 투자원금에는 더하지 않습니다. '
        '투자금(+)과 투자금 회수(-)만 총 투자금액에 반영됩니다.</div>',
        unsafe_allow_html=True
    )

    if st.button(
        "전체 투자원금 기록 저장",
        type="primary",
        width="stretch"
    ):
        saved = 0
        for name, item in pending.items():
            if save_flow(
                name,
                item["date"].isoformat(),
                item["type"],
                item["amount"],
                item["memo"]
            ):
                saved += 1

        if saved:
            st.success(f"{saved}개 계좌의 거래내역을 저장했습니다.")
            st.rerun()
        else:
            st.info("저장할 거래내역이 없습니다.")

elif page == "HLB 투자 메모":
    st.markdown("## 📝 HLB 투자 메모")
    st.caption(
        "그날의 HLB 종가와 매수·매도·홀딩 판단, "
        "개인적인 생각을 기록합니다."
    )

    c1, c2, c3 = st.columns([1,1,1])
    with c1:
        d = st.date_input("날짜", date.today())
    with c2:
        price = money_input("HLB 종가", 0, "hlb_close_price")
    with c3:
        action = st.selectbox(
            "매매 판단",
            ["매수","매도","홀딩"]
        )

    thought = st.text_area(
        "오늘의 생각",
        placeholder="그날 가격과 매수·매도·홀딩 판단의 이유를 기록하세요.",
        height=120
    )

    if st.button(
        "HLB 투자 메모 저장",
        type="primary",
        width="stretch"
    ):
        save_hlb(
            d.isoformat(),
            price,
            action,
            thought
        )
        st.success("HLB 투자 메모를 저장했습니다.")
        st.rerun()

    history = load_hlb()
    if not history.empty:
        st.markdown("### HLB 종가 추이")
        chart = (
            alt.Chart(history)
            .mark_line(point=True)
            .encode(
                x=alt.X(
                    "record_date:T",
                    title=None,
                    axis=alt.Axis(format="%Y.%m", labelAngle=0)
                ),
                y=alt.Y("close_price:Q", title="종가(원)"),
                tooltip=[
                    alt.Tooltip(
                        "record_date:T",
                        title="날짜",
                        format="%Y.%m.%d"
                    ),
                    alt.Tooltip(
                        "close_price:Q",
                        title="HLB 종가",
                        format=",.0f"
                    ),
                    alt.Tooltip("action:N", title="매매 판단")
                ]
            )
            .properties(height=280)
        )
        st.altair_chart(chart, width="stretch")

        show = history.copy()
        show["날짜"] = show["record_date"].dt.strftime("%Y-%m-%d")
        show["HLB 종가"] = show["close_price"].map(
            lambda x: f"{num(x):,.0f}"
        )
        show = show[
            ["날짜","HLB 종가","action","thought"]
        ].rename(columns={
            "action":"매매 판단",
            "thought":"오늘의 생각"
        })
        st.dataframe(
            show,
            width="stretch",
            hide_index=True
        )

elif page == "연금내역":
    st.markdown("## 🧾 연금내역")
    st.caption(
        "좌측에서 새로운 연금자료를 입력하고, "
        "우측에서 직전 기록 대비 변동액·변동률을 확인합니다."
    )

    left, right = st.columns([0.40,0.60], gap="large")

    with left:
        with st.container(border=True):
            st.markdown(
                '<div class="account-title">새 연금 기록</div>',
                unsafe_allow_html=True
            )

            month_text = st.text_input(
                "년월",
                value=date.today().strftime("%Y.%m"),
                placeholder="예: 2026.08"
            )

            lump_sum = money_input(
                "일시금",
                0,
                "pension_lump_sum"
            )
            monthly = money_input(
                "연금(월)",
                0,
                "pension_monthly"
            )
            retirement = money_input(
                "퇴직수당",
                0,
                "pension_retirement"
            )
            memo = st.text_input(
                "메모",
                placeholder="특이사항(선택)"
            )

            if st.button(
                "연금내역 저장",
                type="primary",
                width="stretch"
            ):
                ym = validate_year_month(month_text)
                if not ym:
                    st.error(
                        "년월은 YYYY.MM 형식으로 입력해 주세요. "
                        "예: 2026.08"
                    )
                else:
                    save_pension(
                        ym,
                        lump_sum,
                        monthly,
                        retirement,
                        memo
                    )
                    st.success("연금내역을 저장했습니다.")
                    st.rerun()

    with right:
        st.markdown("### 연금 변동 내역")
        pension_history_view()

elif page == "기록 관리":
    st.markdown("## 🗂️ 기록 관리")

    tab1, tab2, tab3, tab4 = st.tabs(
        [
            "현재가격 기록",
            "투자원금 기록",
            "HLB 투자 메모",
            "연금내역"
        ]
    )

    with tab1:
        ss = load_snapshots()
        if ss.empty:
            st.info("현재가격 기록이 없습니다.")
        else:
            show = ss.copy()
            show["기준일"] = show["record_date"].dt.strftime("%Y-%m-%d")
            show["현재 총 자산"] = show["total_current"].map(kr)
            show["총 투자금액"] = show["total_invested"].map(kr)
            show["수익금"] = show["total_profit"].map(kr)
            show["수익률"] = show["total_return"].map(pc)
            show["등록방식"] = show["source"].map({
                "excel_initial":"초기 이관",
                "manual_current_price":"직접 입력"
            }).fillna(show["source"])

            st.dataframe(
                show[
                    [
                        "id","기준일","총 투자금액","현재 총 자산",
                        "수익금","수익률","등록방식","note"
                    ]
                ].rename(columns={"note":"메모"}),
                width="stretch",
                hide_index=True
            )

    with tab2:
        flows = load_flows()
        if flows.empty:
            st.info("투자원금 기록이 없습니다.")
        else:
            show = flows.copy()
            show["일자"] = show["record_date"].dt.strftime("%Y-%m-%d")
            show["금액"] = show["amount"].map(
                lambda x: f"{num(x):,.0f}"
            )
            show["반영 후 순투자금"] = show["principal_after"].map(
                lambda x: f"{num(x):,.0f}"
            )
            st.dataframe(
                show[
                    [
                        "id","일자","component_name","flow_type",
                        "금액","note","반영 후 순투자금"
                    ]
                ].rename(columns={
                    "component_name":"계좌",
                    "flow_type":"내역",
                    "note":"메모"
                }),
                width="stretch",
                hide_index=True
            )

    with tab3:
        h = load_hlb()
        if h.empty:
            st.info("HLB 투자 메모가 없습니다.")
        else:
            st.dataframe(
                h,
                width="stretch",
                hide_index=True
            )

    with tab4:
        p_hist = load_pension()
        if p_hist.empty:
            st.info("연금내역이 없습니다.")
        else:
            st.dataframe(
                p_hist,
                width="stretch",
                hide_index=True
            )

elif page == "설정/백업":
    st.markdown("## ⚙️ 설정 / 백업")

    st.markdown(
        '<div class="successbox"><b>Excel 초기 이관 완료</b><br>'
        '평소에는 Excel이 필요하지 않으며 모든 새 기록은 '
        'V7 Cloud DB에 저장됩니다.</div>',
        unsafe_allow_html=True
    )

    ss = load_snapshots()
    flows = load_flows()
    hlb = load_hlb()
    pension = load_pension()

    recent = (
        ss["record_date"].max().date().isoformat()
        if not ss.empty else "-"
    )

    st.markdown(
        '<div class="kpi-grid">' +
        card("앱 관리 시작일", sget("initial_date","-"), "초기 이관") +
        card("최근 현재가격일", recent, "최근 저장 기록") +
        card("현재가격 기록 수", f"{len(ss):,}건", "중복 방지") +
        card("투자원금 기록 수", f"{len(flows):,}건", "현금흐름 포함") +
        '</div>',
        unsafe_allow_html=True
    )

    st.markdown("### ☁️ 클라우드 데이터 백업")
    st.caption("현재 Supabase 데이터를 CSV 묶음(ZIP)으로 내려받습니다.")
    def make_cloud_export_zip():
        tables=["settings","component_principals","account_meta","snapshots","component_holdings","account_snapshots","money_flows","trend_history","hlb_memo","pension_history"]
        buffer=io.BytesIO()
        with zipfile.ZipFile(buffer,"w",zipfile.ZIP_DEFLATED) as z:
            for table in tables:
                df=_read_df(f"SELECT * FROM {table}")
                z.writestr(f"{table}.csv",df.to_csv(index=False).encode("utf-8-sig"))
        return buffer.getvalue()
    st.download_button("Supabase 전체 데이터 CSV 백업 다운로드",make_cloud_export_zip(),f"asset_manager_cloud_backup_{date.today().isoformat()}.zip","application/zip",width="stretch")
    st.success("모든 신규 기록은 Supabase PostgreSQL에 저장됩니다. PC가 꺼져 있어도 휴대전화·다른 PC에서 같은 데이터를 사용합니다.")

    st.markdown("### 초기 이관 정보")
    st.write(f"원본 파일: **{sget('initial_file','-')}**")
    st.write(f"최초 이관 기준일: **{sget('initial_date','-')}**")
    st.write(f"HLB 투자 메모: **{len(hlb):,}건**")
    st.write(f"연금내역: **{len(pension):,}건**")

    st.divider()
    change_password_ui()
